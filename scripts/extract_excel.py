"""
Extrae TODO el contenido del Excel de planning del viaje a Japon y genera
src/data/db.json para que la PWA lo consuma sin depender de red.

Uso:
    pip install openpyxl
    python scripts/extract_excel.py

Por defecto lee "JAPO_Planning_Pol-1507 (1).xlsx" en la raiz del proyecto
y escribe "src/data/db.json". Pasa rutas distintas como argumentos:
    python scripts/extract_excel.py "mi_excel.xlsx" "src/data/db.json"
"""

from __future__ import annotations

import datetime
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "item"

ARROW_MARKERS = (" ↗", "↗")  # " ↗" y "↗" sueltos


def clean_text(value):
    """Normaliza el valor de una celda a texto/fecha/numero serializable."""
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value
        for marker in ARROW_MARKERS:
            text = text.replace(marker, "")
        text = text.strip()
        return text or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def cell_link(cell):
    return cell.hyperlink.target if cell.hyperlink else None


def used_max_col(ws: Worksheet, max_row: int) -> int:
    max_col = 1
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for c in row:
            if c.value is not None and c.column > max_col:
                max_col = c.column
    return max_col


def sheet_max_row(ws: Worksheet) -> int:
    max_row = 1
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                max_row = max(max_row, c.row)
    return max_row


def parse_day_outline(ws: Worksheet) -> dict:
    """Convierte la hoja de un dia en un arbol, usando la columna como
    nivel de indentacion (A=nivel0, B=nivel1, C=nivel2, ...)."""
    max_row = sheet_max_row(ws)

    title_cell = ws.cell(row=1, column=1)
    title = clean_text(title_cell.value)

    sections: list[dict] = []
    stack: list[tuple[int, list]] = [(-1, sections)]

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            text = clean_text(cell.value)
            if text is None:
                continue
            depth = cell.column - 1
            link = cell_link(cell)

            while stack and stack[-1][0] >= depth:
                stack.pop()

            node = {"text": text, "children": []}
            if link:
                node["link"] = link

            parent_children = stack[-1][1] if stack else sections
            parent_children.append(node)
            stack.append((depth, node["children"]))

    return {"title": title, "sections": sections}


def parse_index(ws: Worksheet) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=sheet_max_row(ws)):
        day = clean_text(row[0].value)
        date = clean_text(row[1].value)
        if day is None and date is None:
            continue
        rows.append(
            {
                "day": day,
                "date": date,
                "weekday": clean_text(row[2].value),
                "route": clean_text(row[3].value),
            }
        )
    return rows


def parse_resum(ws: Worksheet) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=sheet_max_row(ws)):
        day = clean_text(row[0].value)
        date = clean_text(row[1].value)
        if day is None and date is None:
            continue
        rows.append(
            {
                "day": day,
                "date": date,
                "zone": clean_text(row[2].value),
                "accommodation": clean_text(row[3].value),
                "accommodationLink": cell_link(row[3]),
                "transport": clean_text(row[4].value),
                "distance": clean_text(row[5].value),
            }
        )
    return rows


STATUS_KEYS = {
    "pagat": "paid",
    "reservat (resta pagar)": "reserved",
    "pendent": "pending",
}


def status_key(status: str | None) -> str:
    if status is None:
        return "unknown"
    return STATUS_KEYS.get(status.strip().lower(), "unknown")


# El repo es publico: no exponemos emails personales de los acompanantes.
# Se sustituyen por el mismo nombre que ya aparece en otras filas.
EMAIL_TO_NAME = {
    "apareja4@xtec.cat": "Amaya Pareja de los Santos",
    "sfern121@xtec.cat": "Susana Fernandez",
}


def redact_responsible(value: str | None) -> str | None:
    if value is None:
        return None
    return EMAIL_TO_NAME.get(value.strip().lower(), value)


def parse_reservations(ws: Worksheet) -> dict:
    items = []
    total = None
    legend = None
    seen_ids: dict[str, int] = {}

    for row in ws.iter_rows(min_row=4, max_row=sheet_max_row(ws)):
        concept = clean_text(row[1].value)
        if concept is None:
            continue
        if concept.upper() == "TOTAL":
            total = {
                "costTotal": clean_text(row[3].value),
                "costPerPerson": clean_text(row[4].value),
            }
            continue
        if concept.lower().startswith("llegenda"):
            legend = concept
            continue

        status = clean_text(row[2].value)
        date_val = clean_text(row[0].value)

        # Id estable basado solo en el concepto (no en la fecha ni en la
        # posicion de la fila): la fecha vive a veces en la primera fila de
        # un bloque de varias filas, asi que un insert desplaza que valor a
        # otra fila del mismo bloque aunque el concepto no haya cambiado.
        # El concepto es mucho mas estable ante reordenaciones del Excel.
        base_id = slugify(concept)
        count = seen_ids.get(base_id, 0)
        seen_ids[base_id] = count + 1
        item_id = base_id if count == 0 else f"{base_id}-{count + 1}"

        items.append(
            {
                "id": item_id,
                "date": date_val,
                "concept": concept,
                "link": cell_link(row[1]),
                "status": status,
                "statusKey": status_key(status),
                "costTotal": clean_text(row[3].value),
                "costPerPerson": clean_text(row[4].value),
                "responsible": redact_responsible(clean_text(row[5].value)),
                "notes": clean_text(row[6].value),
                "notesLink": cell_link(row[6]),
                "checkIn": clean_text(row[7].value) if len(row) > 7 else None,
                "checkInLink": cell_link(row[7]) if len(row) > 7 else None,
                "checkOut": clean_text(row[8].value) if len(row) > 8 else None,
                "checkOutLink": cell_link(row[8]) if len(row) > 8 else None,
            }
        )

    return {"items": items, "total": total, "legend": legend}


def main():
    project_root = Path(__file__).resolve().parent.parent

    excel_arg = sys.argv[1] if len(sys.argv) > 1 else None
    output_arg = sys.argv[2] if len(sys.argv) > 2 else None

    if excel_arg:
        excel_path = Path(excel_arg)
    else:
        candidates = sorted(project_root.glob("*.xlsx"))
        if not candidates:
            print("No se encontro ningun .xlsx en la raiz del proyecto.")
            sys.exit(1)
        excel_path = candidates[0]

    output_path = Path(output_arg) if output_arg else project_root / "src" / "data" / "db.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    db = {
        "meta": {
            "generatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
            "sourceFile": excel_path.name,
        },
        "index": parse_index(wb["INDEX"]),
        "resum": parse_resum(wb["RESUM"]),
        "reservations": parse_reservations(wb["RESERVES"]),
        "days": {},
    }

    day_sheets = [name for name in wb.sheetnames if name.upper().startswith("DIA ")]
    day_sheets.sort(key=lambda n: int(n.split(" ")[1]))

    for name in day_sheets:
        day_num = name.split(" ")[1]
        print(f"Procesando {name}...")
        db["days"][day_num] = parse_day_outline(wb[name])

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

    print(f"OK -> {output_path}")
    print(f"Dias extraidos: {len(db['days'])}")
    print(f"Reservas extraidas: {len(db['reservations']['items'])}")


if __name__ == "__main__":
    main()
