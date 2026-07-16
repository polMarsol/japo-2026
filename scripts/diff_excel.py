"""
Compara dos versiones del Excel de planning celda a celda y muestra que
ha cambiado: hojas anadidas/eliminadas, y por cada hoja compartida, que
celdas tienen texto o hyperlink distinto.

Uso:
    python scripts/diff_excel.py "old.xlsx" "new.xlsx"
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return value


def cell_link(cell):
    return cell.hyperlink.target if cell.hyperlink else None


def sheet_max_row_col(ws):
    max_row, max_col = 1, 1
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                max_row = max(max_row, c.row)
                max_col = max(max_col, c.column)
    return max_row, max_col


def diff_sheet(name, ws_old, ws_new):
    changes = []
    max_row = max(sheet_max_row_col(ws_old)[0], sheet_max_row_col(ws_new)[0])
    max_col = max(sheet_max_row_col(ws_old)[1], sheet_max_row_col(ws_new)[1])

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            old_cell = ws_old.cell(row=r, column=c)
            new_cell = ws_new.cell(row=r, column=c)
            old_val, new_val = clean(old_cell.value), clean(new_cell.value)
            old_link, new_link = cell_link(old_cell), cell_link(new_cell)
            if old_val != new_val or old_link != new_link:
                changes.append((r, c, old_val, new_val, old_link, new_link))

    if changes:
        print(f"\n=== {name} ({len(changes)} celdas distintas) ===")
        for r, c, ov, nv, ol, nl in changes:
            col_letter = openpyxl.utils.get_column_letter(c)
            print(f"  {col_letter}{r}:")
            if ov != nv:
                print(f"    texto:  {ov!r}  ->  {nv!r}")
            if ol != nl:
                print(f"    link:   {ol!r}  ->  {nl!r}")


def main():
    if len(sys.argv) != 3:
        print("Uso: python scripts/diff_excel.py old.xlsx new.xlsx")
        sys.exit(1)

    old_path, new_path = Path(sys.argv[1]), Path(sys.argv[2])
    wb_old = openpyxl.load_workbook(old_path, data_only=True)
    wb_new = openpyxl.load_workbook(new_path, data_only=True)

    old_sheets = set(wb_old.sheetnames)
    new_sheets = set(wb_new.sheetnames)

    added = new_sheets - old_sheets
    removed = old_sheets - new_sheets
    common = old_sheets & new_sheets

    if added:
        print(f"Hojas nuevas: {sorted(added)}")
    if removed:
        print(f"Hojas eliminadas: {sorted(removed)}")

    def sheet_sort_key(n):
        parts = n.upper().split(" ")
        if parts[0] == "DIA" and len(parts) > 1 and parts[1].isdigit():
            return (1, int(parts[1]))
        return (0, n)

    for name in sorted(common, key=sheet_sort_key):
        diff_sheet(name, wb_old[name], wb_new[name])

    print("\nOK: diff completo.")


if __name__ == "__main__":
    main()
